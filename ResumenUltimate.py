from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from os import listdir
from statistics import mean, median
import pandas
import re

archivo = 'Resumen.xlsx'
# directorioBrutos = 'C:/Users/Admin/Desktop/Escape/Datos/Brutos/'
# directorioConvertidos = 'C:/Users/Admin/Desktop/Escape/Datos/ConvertidosPython/Escape/'
directorioBrutos = '/home/daniel/Documents/Doctorado/Proyecto de Doctorado/ExperimentoEscape/PruebasBrutos/'
directorioConvertidos = '/home/daniel/Documents/Doctorado/Proyecto de Doctorado/ExperimentoEscape/Flex/'
sesionInicial = 1
sesionFinal = 3

sujetos = ['E3', 'E4', 'E5', 'E7', 'E8', 'E9']
columnasProp = [2, 3, 4, 5, 6, 7, 8]
columnasResp = [2, 9, 16, 23, 30, 37]
columnasLatPal = [2, 7, 12, 17, 22, 27]
columnasEscapes = [2, 13, 24, 35, 46, 57]
columnasLatEsc = [2, 13, 24, 35, 46, 57]


# Convertidor
def convertir(columnas, subfijo):
    for ssn in range(sesionInicial, sesionFinal + 1):
        for sjt in range(len(sujetos)):
            print('Convirtiendo sesión ' + str(ssn) + ' de sujeto ' + sujetos[sjt] + '.')
            # Pandas lee los datos y los escribe en el archivo convertido en 6 columnas separando por los espacios.
            # El argumento names indica cuántas columnas se crearán. Evita errores cuando se edita el archivo de Med.
            datos = pandas.read_csv(directorioBrutos + sujetos[sjt] + subfijo + str(ssn), header=None,
                                    names=range(columnas), sep=r'\s+')
            datos.to_excel(directorioConvertidos + sujetos[sjt] + subfijo + str(ssn) + '.xlsx', index=False,
                           header=None)

            # Openpyxl abre el archivo creado por pandas, lee la hoja y la almacena en la variable hojaCompleta.
            archivoCompleto = load_workbook(directorioConvertidos + sujetos[sjt] + subfijo + str(ssn) + '.xlsx')
            hojaCompleta = archivoCompleto.active

            # Se genera una lista que contenga sub-listas con todos los valores de las listas dadas por Med.
            # Funciona para cualquier cantidad de listas.
            # Los datos se convierten en flotantes para que todos tengan punto decimal, y luego en string para que mas
            # adelante el método split los pueda separar por el punto.
            metalista = [[]]
            contadormetalista = 0

            columna1 = hojaCompleta['B']
            for fila in range(12, len(columna1)):
                for columna in range(2, columnas + 1):
                    if hojaCompleta[get_column_letter(columna) + str(fila)].value is not None:
                        metalista[contadormetalista].append(str(float(hojaCompleta[get_column_letter(columna) +
                                                                                   str(fila)].value)))
                    elif hojaCompleta[get_column_letter(columna) + str(fila)].value is None and columna == 2:
                        metalista.append([])
                        contadormetalista += 1

            # Escribir cada sub-lista en una columna de excel separando por punto decimal.
            # Se utilizan expresiones regulares (regex) para indicar al programa que debe añadir ceros cuando pandas
            # los ha eliminado (cuando están al final de una cifra después de un punto decimal).
            regex1 = re.compile(r'^\d+\.\d{2}$')
            for i in range(len(metalista)):
                for j in range(len(metalista[i])):
                    if regex1.search(metalista[i][j]):
                        metalista[i][j] += '0'
                    hojaCompleta[get_column_letter((i * 2) + 9) + str(j + 1)] = int(metalista[i][j].split('.')[0])
                    hojaCompleta[get_column_letter((i * 2) + 10) + str(j + 1)] = int(metalista[i][j].split('.')[1])
            archivoCompleto.save(directorioConvertidos + sujetos[sjt] + subfijo + str(ssn) + '.xlsx')
        print('\n')


# Función para crear hojas.
def hoja(nombre):
    if nombre not in wb.sheetnames:
        return wb.create_sheet(nombre)
    else:
        return wb[nombre]


# Función para contar respuestas por tipo de ensayo. Los argumentos son marcadores de Med.
def conteoresp(inicioEnsayo, finEnsayo, respuesta):
    contadorTemp = 0
    inicio = 0
    resp = []
    for n in range(1, len(marcadores)):
        if marcadores[n].value == inicioEnsayo:
            inicio = 1
        elif marcadores[n].value == respuesta and inicio == 1:
            contadorTemp += 1
        elif marcadores[n].value == finEnsayo and inicio == 1:
            inicio = 0
            resp.append(contadorTemp)
            contadorTemp = 0
    return resp


# Función para contar respuestas totales. El argumento es el marcador de Med.
def conteototal(respuesta):
    contador = 0
    for n in range(len(marcadores)):
        if marcadores[n].value == respuesta:
            contador += 1
    return contador


# Función para contar latencias. Si en un ensayo no hay respuestas que contar, la función resulta en una lista con un
# cero. Los argumentos son marcadores de Med.
def conteolat(inicioensayo, respuesta):
    inicio = 0
    lat = []
    tiempoini = 0
    for n in range(1, len(marcadores)):
        if marcadores[n].value == inicioensayo:
            inicio = 1
            tiempoini = tiempo[n].value
        elif marcadores[n].value == respuesta and inicio == 1:
            lat.append((tiempo[n].value - tiempoini) / 20)
            inicio = 0
    if len(lat) == 0:
        lat = [0]
    return lat


# Función para escribir listas en columnas. El argumento "restar" indica si se debe restar uno a las respuestas dadas
# en cada ensayo. Esto solo sucede en las palancas dado que se registra también la respuesta que le da inicio al ensayo.
def esccolumnas(titulo, columna, lista, restar):
    hojaind[get_column_letter(columna) + str(1)] = titulo
    if restar:
        for pos in range(len(lista)):
            hojaind[get_column_letter(columna) + str(pos + 2)] = lista[pos] - 1
    else:
        for pos in range(len(lista)):
            hojaind[get_column_letter(columna) + str(pos + 2)] = lista[pos]


convertir(6, '_LIBRES_')

# Resumen
# Revisar si el archivo de resumen ya existe. De lo contrario, crearlo.
if archivo in listdir(directorioConvertidos):
    print('Archivo encontrado. Abriendo...')
    wb = load_workbook(directorioConvertidos + archivo)
else:
    print('Archivo no encontrado. Creando...')
    wb = Workbook()

# Crear todas las hojas.
proporciones = hoja('Proporciones')
respuestas = hoja('Respuestas')
latencias = hoja('Latencias')
comedero = hoja('Comedero')
escapes = hoja('Escapes')
latNosepoke = hoja('LatNosepoke')

# Loop principal.
for sesion in range(sesionInicial, sesionFinal + 1):
    print('\nIntentando sesión ' + str(sesion) + '...')
    for sujeto in range(len(sujetos)):
        print('Intentando sujeto ' + sujetos[sujeto] + '...')
        sujetoWb = load_workbook(directorioConvertidos + sujetos[sujeto] + '_LIBRES_' + str(sesion) + '.xlsx')
        sujetoWs = sujetoWb.worksheets[0]
        tiempo = sujetoWs['O']
        marcadores = sujetoWs['P']

        # Abrir o crear la hoja para pegar respuestas individuales.
        if 'Respuestas por ensayo' not in sujetoWb.sheetnames:
            hojaind = sujetoWb.create_sheet('Respuestas por ensayo')
        else:
            hojaind = sujetoWb['Respuestas por ensayo']

        # Proporciones
        proporciones[get_column_letter(columnasProp[sujeto]) + str(sesion + 3)] = sujetoWs.cell(14, 6).value
        # Se toma el valor de la celda F14, que corresponde a los valores de fila 14 y columna 6.

        # Conteo de respuestas en palancas en ensayos forzados.
        # Se escriben las respuestas por ensayo en el archivo individual.
        # Se resta una unidad a las medias debido a que conteoresp() cuenta también la respuesta que inicia el ensayo.
        resPalForzDiscRef = conteoresp(114, 180, 202)
        mediaResPalForzDiscRef = mean(resPalForzDiscRef) - 1
        esccolumnas('PalForzDiscRef', 1, resPalForzDiscRef, True)
        # 114: Inicio TL Forz Disc Ref  //  180: Fin ensayo  //  202: Resp Pal Disc
        resPalForzDiscNoRef = conteoresp(115, 180, 202)
        mediaResPalForzDiscNoRef = mean(resPalForzDiscNoRef) - 1
        esccolumnas('PalForzDiscNoRef', 3, resPalForzDiscNoRef, True)
        # 115: Inicio TL Forz Disc NoRef  //  180: Fin ensayo (por TF)  //  202: Res Pal Disc
        resPalForzNoDisc1 = conteoresp(134, 180, 201)
        mediaResPalForzNoDisc1 = mean(resPalForzNoDisc1) - 1
        esccolumnas('PalForzNoDisc1', 5, resPalForzNoDisc1, True)
        # 134: Inicio TL Forz NoDisc 1  //  180: Fin ensayo  //  201: Resp Pal NoDisc
        resPalForzNoDisc2 = conteoresp(137, 180, 201)
        mediaResPalForzNoDisc2 = mean(resPalForzNoDisc2) - 1
        esccolumnas('PalForzNoDisc2', 7, resPalForzNoDisc2, True)
        # 137: Inicio TL Forz NoDisc 2  //  180: Fin ensayo  //  201: Resp Pal NoDisc
        respuestas[get_column_letter(columnasResp[sujeto]) + str(sesion + 3)] = mediaResPalForzDiscRef
        respuestas[get_column_letter(columnasResp[sujeto] + 1) + str(sesion + 3)] = mediaResPalForzDiscNoRef
        respuestas[get_column_letter(columnasResp[sujeto] + 2) + str(sesion + 3)] = mediaResPalForzNoDisc1
        respuestas[get_column_letter(columnasResp[sujeto] + 3) + str(sesion + 3)] = mediaResPalForzNoDisc2

        # Conteo de latencias a palancas en eslabones iniciales.
        # Las latencias por ensayo se pegan en el archivo individual.
        latPalDisc = conteolat(112, 113)
        medianaLatPalDisc = median(latPalDisc)
        esccolumnas('LatPalDisc', 9, latPalDisc, False)
        # 112: IL Forz Disc  //  113: Res Pal
        latPalNoDisc = conteolat(132, 133)
        medianaLatPalNoDisc = median(conteolat(132, 133))
        esccolumnas('LatPalNoDisc', 11, latPalNoDisc, False)
        # 132: IL Forzado NoDisc  //  133: Res Pal
        latencias[get_column_letter(columnasLatPal[sujeto]) + str(sesion + 3)] = medianaLatPalDisc
        latencias[get_column_letter(columnasLatPal[sujeto] + 1) + str(sesion + 3)] = medianaLatPalNoDisc

        # Conteo de respuestas en comederos en ensayos forzados
        # Las respuestas por ensayo se pegan en el archivo individual.
        resComForzDiscRef = conteoresp(114, 16, 203)
        mediaResComForzDiscRef = mean(resComForzDiscRef)
        esccolumnas('ComForzDiscRef', 13, resComForzDiscRef, False)
        # 114: Inicio TL Forz Disc Ref  //  16: Fin ensayo  //  203: Res Com
        resComForzDiscNoRef = conteoresp(115, 117, 203)
        mediaResComForzDiscNoRef = mean(resComForzDiscNoRef)
        esccolumnas('ComForzDiscNoRef', 15, resComForzDiscNoRef, False)
        # 115: Inicio TL Forz Disc NoRef  //  117: Fin ensayo (por TF)  //  203: Res Com
        resComForzNoDisc1 = conteoresp(134, 40, 203)
        mediaResComForzNoDisc1 = mean(resComForzNoDisc1)
        esccolumnas('ComForzNoDisc1', 17, resComForzNoDisc1, False)
        # 134: Inicio TL Forz NoDisc 1  //  40: Fin ensayo  //  203: Res Com
        resComForzNoDisc2 = conteoresp(137, 43, 203)
        mediaResComForzNoDisc2 = mean(resComForzNoDisc2)
        esccolumnas('ComForzNoDisc2', 19, resComForzNoDisc2, False)
        # 137: Inicio TL Forz NoDisc 2  //  43: Fin ensayo  //  203: Res Com
        comedero[get_column_letter(columnasResp[sujeto]) + str(sesion + 3)] = mediaResComForzDiscRef
        comedero[get_column_letter(columnasResp[sujeto] + 1) + str(sesion + 3)] = mediaResComForzDiscNoRef
        comedero[get_column_letter(columnasResp[sujeto] + 2) + str(sesion + 3)] = mediaResComForzNoDisc1
        comedero[get_column_letter(columnasResp[sujeto] + 3) + str(sesion + 3)] = mediaResComForzNoDisc2

        # Conteo de respuestas en nosepoke.
        escapeForzDiscRef = conteototal(301)
        escapeForzDiscNoRef = conteototal(302)
        escapeForzNoDisc1 = conteototal(303)
        escapeForzNoDisc2 = conteototal(304)
        escapeLibDiscRef = conteototal(305)
        escapeLibDiscNoRef = conteototal(306)
        escapeLibNoDisc1 = conteototal(307)
        escapeLibNoDisc2 = conteototal(308)
        # 301 - 308: Respuestas nosepoke por tipo de ensayo.
        escapes[get_column_letter(columnasEscapes[sujeto]) + str(sesion + 3)] = escapeForzDiscRef
        escapes[get_column_letter(columnasEscapes[sujeto] + 1) + str(sesion + 3)] = escapeForzDiscNoRef
        escapes[get_column_letter(columnasEscapes[sujeto] + 2) + str(sesion + 3)] = escapeForzNoDisc1
        escapes[get_column_letter(columnasEscapes[sujeto] + 3) + str(sesion + 3)] = escapeForzNoDisc2
        escapes[get_column_letter(columnasEscapes[sujeto] + 4) + str(sesion + 3)] = escapeLibDiscRef
        escapes[get_column_letter(columnasEscapes[sujeto] + 5) + str(sesion + 3)] = escapeLibDiscNoRef
        escapes[get_column_letter(columnasEscapes[sujeto] + 6) + str(sesion + 3)] = escapeLibNoDisc1
        escapes[get_column_letter(columnasEscapes[sujeto] + 7) + str(sesion + 3)] = escapeLibNoDisc2

        # Conteo de latencias a nosepoke.
        # Las latencias por ensayo se pegan en el archivo individual.
        latEscForzDiscRef = conteolat(114, 301)
        medianaLatEscForzDiscRef = median(latEscForzDiscRef)
        esccolumnas('LatEscForzDiscRef', 21, latEscForzDiscRef, False)

        latEscForzDiscNoRef = conteolat(115, 302)
        medianaLatEscForzDiscNoRef = median(latEscForzDiscNoRef)
        esccolumnas('LatEscForzDiscNoRef', 23, latEscForzDiscNoRef, False)

        latEscForzNoDisc1 = conteolat(134, 303)
        medianaLatEscForzNoDisc1 = median(latEscForzNoDisc1)
        esccolumnas('LatEscForzNoDisc1', 25, latEscForzNoDisc1, False)

        latEscForzNoDisc2 = conteolat(137, 304)
        medianaLatEscForzNoDisc2 = median(latEscForzNoDisc2)
        esccolumnas('LatEscForzNoDisc2', 27, latEscForzNoDisc2, False)

        latEscLibDiscRef = conteolat(154, 305)
        medianaLatEscLibDiscRef = median(latEscLibDiscRef)
        esccolumnas('LatEscLibDiscRef', 29, latEscLibDiscRef, False)

        latEscLibDiscNoRef = conteolat(155, 306)
        medianaLatEscLibDiscNoRef = median(latEscLibDiscNoRef)
        esccolumnas('LatEscLibDiscNoRef', 31, latEscLibDiscNoRef, False)

        latEscLibNoDisc1 = conteolat(157, 307)
        medianaLatEscLibNoDisc1 = median(latEscLibNoDisc1)
        esccolumnas('LatEscLibNoDisc1', 33, latEscLibNoDisc1, False)

        latEscLibNoDisc2 = conteolat(160, 308)
        medianaLatEscLibNoDisc2 = median(latEscLibNoDisc2)
        esccolumnas('LatEscLibNoDisc2', 35, latEscLibNoDisc2, False)
        # El primer marcador es el inicio de su tipo de ensayo; el segundo, la respuesta de escape correspondiente.

        latNosepoke[get_column_letter(columnasEscapes[sujeto]) + str(sesion + 3)] = medianaLatEscForzDiscRef
        latNosepoke[get_column_letter(columnasEscapes[sujeto] + 1) + str(sesion + 3)] = medianaLatEscForzDiscNoRef
        latNosepoke[get_column_letter(columnasEscapes[sujeto] + 2) + str(sesion + 3)] = medianaLatEscForzNoDisc1
        latNosepoke[get_column_letter(columnasEscapes[sujeto] + 3) + str(sesion + 3)] = medianaLatEscForzNoDisc2
        latNosepoke[get_column_letter(columnasEscapes[sujeto] + 4) + str(sesion + 3)] = medianaLatEscLibDiscRef
        latNosepoke[get_column_letter(columnasEscapes[sujeto] + 5) + str(sesion + 3)] = medianaLatEscLibDiscNoRef
        latNosepoke[get_column_letter(columnasEscapes[sujeto] + 6) + str(sesion + 3)] = medianaLatEscLibNoDisc1
        latNosepoke[get_column_letter(columnasEscapes[sujeto] + 7) + str(sesion + 3)] = medianaLatEscLibNoDisc2

        sujetoWb.save(directorioConvertidos + sujetos[sujeto] + '_LIBRES_' + str(sesion) + '.xlsx')

wb.save(directorioConvertidos + archivo)
