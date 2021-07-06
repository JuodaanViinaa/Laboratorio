from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from xlrd import open_workbook
from os import listdir
from statistics import mean, median

archivo = 'Resumen.xlsx'
directorio = '/home/daniel/Documents/Doctorado/Proyecto de Doctorado/ExperimentoEscape/Datos/'
sesionInicial = 1
sesionFinal = 3

sujetos = ['E3', 'E4', 'E5', 'E7', 'E8', 'E9']
columnasProp = [2, 3, 4, 5, 6, 7, 8]
columnasResp = [2, 9, 16, 23, 30, 37]
columnasLatPal = [2, 7, 12, 17, 22, 27]
columnasEscapes = [2, 13, 24, 35, 46, 57]
columnasLatEsc = [2, 13, 24, 35, 46, 57]

# Revisar si el archivo ya existe. De lo contrario, crearlo.
if archivo in listdir(directorio):
    print('Archivo encontrado. Abriendo...')
    wb = load_workbook(directorio + archivo)
else:
    print('Archivo no encontrado. Creando...')
    wb = Workbook()


# Función para crear hojas.
def hoja(nombre):
    if nombre not in wb.sheetnames:
        return wb.create_sheet(nombre)
    else:
        return wb[nombre]


# Función para contar respuestas por tipo de ensayo. Los argumentos son marcadores de Med.
def conteoresp(inicioEnsayo, finEnsayo, respuesta):
    contadorTemp = 0
    inicio = 0
    resp = []
    for n in range(1, len(marcadores)):
        if marcadores[n] == inicioEnsayo:
            inicio = 1
        elif marcadores[n] == respuesta and inicio == 1:
            contadorTemp += 1
        elif marcadores[n] == finEnsayo and inicio == 1:
            inicio = 0
            resp.append(contadorTemp)
            contadorTemp = 0
    return resp


# Función para contar respuestas totales. El argumento es el marcador de Med.
def conteototal(respuesta):
    contador = 0
    for n in range(len(marcadores)):
        if marcadores[n] == respuesta:
            contador += 1
    return contador


# Función para contar latencias. Si en un ensayo no hay respuestas que contar, la función resulta en una lista con un
#  cero. Los argumentos son marcadores de Med.
def conteolat(inicioensayo, respuesta):
    inicio = 0
    lat = []
    tiempoini = 0
    for n in range(1, len(marcadores)):
        if marcadores[n] == inicioensayo:
            inicio = 1
            tiempoini = tiempo[n]
        elif marcadores[n] == respuesta and inicio == 1:
            lat.append((tiempo[n] - tiempoini) / 20)
            inicio = 0
    if len(lat) == 0:
        lat = [0]
    return lat


# Crear todas las hojas.
proporciones = hoja('Proporciones')
respuestas = hoja('Respuestas')
latencias = hoja('Latencias')
comedero = hoja('Comedero')
escapes = hoja('Escapes')
latNosepoke = hoja('LatNosepoke')

# Loop principal.
for sesion in range(sesionInicial, sesionFinal + 1):
    print('\nIntentando sesión ' + str(sesion) + '...')
    for sujeto in range(len(sujetos)):
        print('Intentando sujeto ' + sujetos[sujeto] + '...')
        sujetoWb = open_workbook(directorio + sujetos[sujeto] + '_LIBRES_' + str(sesion) + '.xls')
        sujetoWs = sujetoWb.sheet_by_index(0)
        tiempo = sujetoWs.col_values(15)
        marcadores = sujetoWs.col_values(16)

        # Proporciones
        proporciones[get_column_letter(columnasProp[sujeto]) + str(sesion + 3)] = sujetoWs.cell_value(16, 6)
        # Se toma el valor de la celda G17, que corresponde a los valores de fila 16 y columna 6 debido a que
        # el conteo comienza desde 0 y no desde 1.

        # Conteo de respuestas en palancas en ensayos forzados.
        mediaResPalForzDiscRef = mean(conteoresp(114, 180, 202)) - 1
        # 114: Inicio TL Forz Disc Ref  //  180: Fin ensayo  //  202: Resp Pal Disc
        mediaResPalForzDiscNoRef = mean(conteoresp(115, 180, 202)) - 1
        # 115: Inicio TL Forz Disc NoRef  //  180: Fin ensayo (por TF)  //  202: Res Pal Disc
        # Se restaba 1 debido a que este tipo de ensayo termina por tiempo y no por respuesta, y el programa
        # cuenta una respuesta adicional por defecto (la respuesta que da inicio al TL).
        mediaResPalForzNoDisc1 = mean(conteoresp(134, 180, 201)) - 1
        # 134: Inicio TL Forz NoDisc 1  //  180: Fin ensayo  //  201: Resp Pal NoDisc
        mediaResPalForzNoDisc2 = mean(conteoresp(137, 180, 201)) - 1
        # 137: Inicio TL Forz NoDisc 2  //  180: Fin ensayo  //  201: Resp Pal NoDisc
        respuestas[get_column_letter(columnasResp[sujeto]) + str(sesion + 3)] = mediaResPalForzDiscRef
        respuestas[get_column_letter(columnasResp[sujeto] + 1) + str(sesion + 3)] = mediaResPalForzDiscNoRef
        respuestas[get_column_letter(columnasResp[sujeto] + 2) + str(sesion + 3)] = mediaResPalForzNoDisc1
        respuestas[get_column_letter(columnasResp[sujeto] + 3) + str(sesion + 3)] = mediaResPalForzNoDisc2

        # Conteo de latencias a palancas en eslabones iniciales.
        medianaLatPalDisc = median(conteolat(112, 113))
        # 112: IL Forz Disc  //  113: Res Pal
        medianaLatPalNoDisc = median(conteolat(132, 133))
        # 132: IL Forzado NoDisc  //  133: Res Pal
        latencias[get_column_letter(columnasLatPal[sujeto]) + str(sesion + 3)] = medianaLatPalDisc
        latencias[get_column_letter(columnasLatPal[sujeto] + 1) + str(sesion + 3)] = medianaLatPalNoDisc

        # Conteo de respuestas en comederos en ensayos forzados
        mediaResComForzDiscRef = mean(conteoresp(114, 16, 203))
        # 114: Inicio TL Forz Disc Ref  //  16: Fin ensayo  //  203: Res Com
        mediaResComForzDiscNoRef = mean(conteoresp(115, 117, 203))
        # 115: Inicio TL Forz Disc NoRef  //  117: Fin ensayo (por TF)  //  203: Res Com
        mediaResComForzNoDisc1 = mean(conteoresp(134, 40, 203))
        # 134: Inicio TL Forz NoDisc 1  //  40: Fin ensayo  //  203: Res Com
        mediaResComForzNoDisc2 = mean(conteoresp(137, 43, 203))
        # 137: Inicio TL Forz NoDisc 2  //  43: Fin ensayo  //  203: Res Com
        comedero[get_column_letter(columnasResp[sujeto]) + str(sesion + 3)] = mediaResComForzDiscRef
        comedero[get_column_letter(columnasResp[sujeto] + 1) + str(sesion + 3)] = mediaResComForzDiscNoRef
        comedero[get_column_letter(columnasResp[sujeto] + 2) + str(sesion + 3)] = mediaResComForzNoDisc1
        comedero[get_column_letter(columnasResp[sujeto] + 3) + str(sesion + 3)] = mediaResComForzNoDisc2

        # Conteo de respuestas en nosepoke.
        escapeForzDiscRef = conteototal(301)
        escapeForzDiscNoRef = conteototal(302)
        escapeForzNoDisc1 = conteototal(303)
        escapeForzNoDisc2 = conteototal(304)
        escapeLibDiscRef = conteototal(305)
        escapeLibDiscNoRef = conteototal(306)
        escapeLibNoDisc1 = conteototal(307)
        escapeLibNoDisc2 = conteototal(308)
        # 301 - 308: Respuestas nosepoke por tipo de ensayo.
        escapes[get_column_letter(columnasEscapes[sujeto]) + str(sesion + 3)] = escapeForzDiscRef
        escapes[get_column_letter(columnasEscapes[sujeto] + 1) + str(sesion + 3)] = escapeForzDiscNoRef
        escapes[get_column_letter(columnasEscapes[sujeto] + 2) + str(sesion + 3)] = escapeForzNoDisc1
        escapes[get_column_letter(columnasEscapes[sujeto] + 3) + str(sesion + 3)] = escapeForzNoDisc2
        escapes[get_column_letter(columnasEscapes[sujeto] + 4) + str(sesion + 3)] = escapeLibDiscRef
        escapes[get_column_letter(columnasEscapes[sujeto] + 5) + str(sesion + 3)] = escapeLibDiscNoRef
        escapes[get_column_letter(columnasEscapes[sujeto] + 6) + str(sesion + 3)] = escapeLibNoDisc1
        escapes[get_column_letter(columnasEscapes[sujeto] + 7) + str(sesion + 3)] = escapeLibNoDisc2

        # Conteo de latencias a nosepoke.
        medianaLatEscForzDiscRef = median(conteolat(114, 301))
        medianaLatEscForzDiscNoRef = median(conteolat(115, 302))
        medianaLatEscForzNoDisc1 = median(conteolat(134, 303))
        medianaLatEscForzNoDisc2 = median(conteolat(137, 304))
        medianaLatEscLibDiscRef = median(conteolat(154, 305))
        medianaLatEscLibDiscNoRef = median(conteolat(155, 306))
        medianaLatEscLibNoDisc1 = median(conteolat(157, 307))
        medianaLatEscLibNoDisc2 = median(conteolat(160, 308))
        # El primer marcador es el inicio de su tipo de ensayo; el segundo, la respuesta de escape correspondiente.

        latNosepoke[get_column_letter(columnasEscapes[sujeto]) + str(sesion + 3)] = medianaLatEscForzDiscRef
        latNosepoke[get_column_letter(columnasEscapes[sujeto] + 1) + str(sesion + 3)] = medianaLatEscForzDiscNoRef
        latNosepoke[get_column_letter(columnasEscapes[sujeto] + 2) + str(sesion + 3)] = medianaLatEscForzNoDisc1
        latNosepoke[get_column_letter(columnasEscapes[sujeto] + 3) + str(sesion + 3)] = medianaLatEscForzNoDisc2
        latNosepoke[get_column_letter(columnasEscapes[sujeto] + 4) + str(sesion + 3)] = medianaLatEscLibDiscRef
        latNosepoke[get_column_letter(columnasEscapes[sujeto] + 5) + str(sesion + 3)] = medianaLatEscLibDiscNoRef
        latNosepoke[get_column_letter(columnasEscapes[sujeto] + 6) + str(sesion + 3)] = medianaLatEscLibNoDisc1
        latNosepoke[get_column_letter(columnasEscapes[sujeto] + 7) + str(sesion + 3)] = medianaLatEscLibNoDisc2

wb.save(directorio + archivo)
