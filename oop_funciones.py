from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from os import listdir
from statistics import mean, median
import pandas
import re
from shutil import move


def fetch(sheet, origin_cell_row, origin_cell_column):
    """
    La función regresa el valor contenido dentro de una celda en una hoja de cálculo especificada.\n
    :param sheet: La hoja de cálculo leída.
    :param origin_cell_column: Columna en que se encuentra la celda.
    :param origin_cell_row: Fila en que se encuentra la celda.
    :return: Valor de la celda referida.
    """
    return sheet.cell(origin_cell_row, origin_cell_column).value


def conteoresp(marks, trialStart, trialEnd, response):  # Count_per_trial
    """
    Cuenta respuestas entre el marcador de inicio de ensayo y el de fin de ensayo. En caso de que el marcador de la
    respuesta a contar sea el mismo que el de la respuesta que da inicio al ensayo se contará una respuesta adicional.
    Esto puede corregirse en el análisis principal con la opción "substract", y en la función 'esccolumnas' mediante el
    parámetro "restar".\n
    :param marks: Lista con los marcadores.
    :param trialStart: Marcador de inicio de ensayo.
    :param trialEnd: Marcador de fin de ensayo.
    :param response: Marcador de respuesta a contar.
    :return: Lista con la cantidad de respuestas ocurridas por ensayo.
    """
    contadorTemp = 0
    inicio = 0
    resp = []
    for n in range(1, len(marks)):
        if marks[n].value == trialStart:
            inicio = 1
        elif marks[n].value == response and inicio == 1:
            contadorTemp += 1
        elif marks[n].value == trialEnd and inicio == 1:
            inicio = 0
            resp.append(contadorTemp)
            contadorTemp = 0
    if len(resp) == 0:
        resp = [0]
    return resp


def resp_dist(marks, time, trialStart, trialEnd, response, bin_size, bin_amount):
    """
    Cuenta las respuestas por bin de tiempo dentro de cada ensayo de la sesión. Se puede elegir la cantidad de bins y su
    tamaño. La función puede lidiar tanto con situaciones en las que hay marcador de fin de ensayo (es decir, hay
    intervalo entre ensayos) como situaciones en las que no. En caso de que el ensayo continúe más allá del último bin
    declarado se contabilizarán todas las respuestas dadas desde el fin del último bin declarado hasta el fin del ensayo
    como un único gran bin.\n
    :param marks: Lista con los marcadores.
    :param time: Lista con el tiempo de la sesión.
    :param trialStart: Marcador de inicio de ensayo.
    :param trialEnd: Marcador de fin de ensayo.
    :param response: Marcador de respuesta a contar.
    :param bin_size: Tamaño en segundos de los bins.
    :param bin_amount: Cantidad de bins por ensayo.
    :return: Lista compuesta de sublistas con las respuestas ocurridas por bin por ensayo.
    """
    inicio = 0
    resp_por_ensayo = [0] * (bin_amount + 1)  # Generar lista con tantos ceros como diga el parámetro bin_amount
    resp_totales = []
    bin_tuples = []
    if trialStart == trialEnd:
        for index, mark in enumerate(marks):
            if mark.value == trialStart and inicio == 0:
                tiempo_inicio = time[index].value
                # Este loop crea una lista con tantas tuplas como bin_amount dicte. Cada tupla contendrá el tiempo de inicio
                # y de fin de cada bin. El tiempo de fin de un bin es igual al tiempo de inicio del siguiente.
                for i in range(bin_amount):
                    tiempo_fin = tiempo_inicio + (bin_size * 20)
                    bin_tuples.append((tiempo_inicio, tiempo_fin))
                    tiempo_inicio = tiempo_fin
                inicio = 1

            elif mark.value == trialStart and inicio == 1:
                resp_totales.append(resp_por_ensayo)
                resp_por_ensayo = [0] * (bin_amount + 1)
                bin_tuples = []
                tiempo_inicio = time[index].value
                for i in range(bin_amount):
                    tiempo_fin = tiempo_inicio + (bin_size * 20)
                    bin_tuples.append((tiempo_inicio, tiempo_fin))
                    tiempo_inicio = tiempo_fin

            elif mark.value == response and inicio == 1:
                for idx, bin_tuple in enumerate(bin_tuples):
                    if bin_tuple[0] <= time[index].value < bin_tuple[1]:
                        # Si se encuentra una respuesta, se comienza a ciclar a través de la lista de tuplas. Si el tiempo
                        # en que ocurrió la respuesta está contenido en alguno de los intervalos dictados por las tuplas, se
                        # agrega una unidad a la lista de resp_por_ensayo en la misma posición que tenga la tupla en su
                        # propia lista.
                        resp_por_ensayo[idx] += 1
                    elif time[index].value >= bin_tuples[-1][-1]:
                        # Si el tiempo está más allá del dictado por las tuplas se agrega una unidad a la última posición de
                        # la lista resp_por_ensayo. Esta última posición contendrá el aglomerado de todas las respuestas que
                        # ocurran después de los esperado según el parámetro bin_amount.
                        resp_por_ensayo[-1] += 1
        resp_totales.append(resp_por_ensayo)

    else:
        for index, mark in enumerate(marks):
            if mark.value == trialStart and inicio == 0:
                tiempo_inicio = time[index].value
                # Este loop crea una lista con tantas tuplas como bin_amount dicte. Cada tupla contendrá el tiempo de inicio
                # y de fin de cada bin. El tiempo de fin de un bin es igual al tiempo de inicio del siguiente.
                for i in range(bin_amount):
                    tiempo_fin = tiempo_inicio + (bin_size * 20)
                    bin_tuples.append((tiempo_inicio, tiempo_fin))
                    tiempo_inicio = tiempo_fin
                inicio = 1

            elif mark.value == response and inicio == 1:
                for idx, bin_tuple in enumerate(bin_tuples):
                    if bin_tuple[0] <= time[index].value < bin_tuple[1]:
                        # Si se encuentra una respuesta, se comienza a ciclar a través de la lista de tuplas. Si el tiempo
                        # en que ocurrió la respuesta está contenido en alguno de los intervalos dictados por las tuplas, se
                        # agrega una unidad a la lista de resp_por_ensayo en la misma posición que tenga la tupla en su
                        # propia lista.
                        resp_por_ensayo[idx] += 1
                    elif time[index].value >= bin_tuples[-1][-1]:
                        # Si el tiempo está más allá del dictado por las tuplas se agrega una unidad a la última posición de
                        # la lista resp_por_ensayo. Esta última posición contendrá el aglomerado de todas las respuestas que
                        # ocurran después de los esperado según el parámetro bin_amount.
                        resp_por_ensayo[-1] += 1

            elif mark.value == trialEnd:
                # Al finalizar cada ensayo la lista con las respuestas por ensayo se agrega a una lista de orden superior
                # llamada resp_totales. Los valores de resp_por_ensayo y bin_tuples se reinician.
                inicio = 0
                resp_totales.append(resp_por_ensayo)
                resp_por_ensayo = [0] * (bin_amount + 1)
                bin_tuples = []

    return resp_totales


def conteototal(marks, response):
    """
    Cuenta la cantidad total de veces que ocurrió un marcador particular en la sesión independientemente de los ensayos.\n
    :param marks: Lista con los marcadores.
    :param response: Marcador de respuesta a contar.
    :return: Integer con la cantidad de ocasiones que ocurrió una respuesta.
    """
    contador = 0
    for n in range(len(marks)):
        if marks[n].value == response:
            contador += 1
    return contador


def conteolat(marks, time, trialStart, response):
    """
    Cuenta la latencia entre el inicio de un ensayo y una respuesta.\n
    :param marks: Lista con los marcadores.
    :param time: Lista con el tiempo de la sesión.
    :param trialStart: Marcador de inicio de ensayo.
    :param response: Marcador de respuesta de interés.
    :return: Lista con las latencias por ensayo.
    """
    inicio = 0
    lat = []
    tiempoini = 0
    for n in range(1, len(marks)):
        if marks[n].value == trialStart:
            inicio = 1
            tiempoini = time[n].value
        elif marks[n].value == response and inicio == 1:
            lat.append((time[n].value - tiempoini) / 20)
            inicio = 0
    if len(lat) == 0:
        lat = [0]
    return lat


def esccolumnas(indivSheet, header, column, data):
    """
    Escribe listas completas en columnas. Útil para escribir los datos completos en los archivos individuales.\n
    :param indivSheet: Objeto de tipo Worksheet (Openpyxl) en el que se escribirá la lista.
    :param header: Rótulo que se escribirá en la primera celda de la columna.
    :param column: Número de la columna en que se escribirá la lista (1-A, 2-B, etc.).
    :param data: Lista a escribir en la columna.
    """
    indivSheet[get_column_letter(column) + str(1)] = header
    for pos in range(len(data)):
        indivSheet[get_column_letter(column) + str(pos + 2)] = data[pos]


def template():
    """
    Imprime un formato para facilitar la declaración de los análisis.
    """
    print("""
    analysis_list = [
    {"fetch": {"cell_row": 10,
               "cell_column": 10,
               "sheet": "Sheet_1",
               "summary_column_list": column_dictionary,
               "offset": 0
               }},

    {"conteoresp": {"measures": 2, # Optional argument. Default value: 1
                    "inicio_ensayo": 111, "fin_ensayo": 222, "respuesta": 333,
                    "inicio_ensayo2": 444, "fin_ensayo2": 555, "respuesta2": 666, # Optional marks. Depends on the value of "measures"
                    "substract": True, # Optional argument. Default value: False
                    "column": 1,
                    "header": "Generic_title",
                    "sheet": "Sheet_2",
                    "summary_column_list": column_dictionary2,
                    "offset": 0,
                    }},

    {"conteototal": {"measures": 2, # Optional argument. Default value: 1
                     "respuesta": 111,
                     "respuesta2": 222, # Optional mark. Depends on the value of "measures"
                     "column": 3,
                     "header": "Generic_title",
                     "sheet": "Sheet_4",
                     "summary_column_list": column_dictionary4,
                     "offset": 0,
                     }},

    {"conteolat": {"measures": 2, # Optional argument. Default value: 1
                   "inicio_ensayo": 111, "respuesta": 222,
                   "inicio_ensayo2": 333, "respuesta2": 444, # Optional marks. Depends on the value of "measures"
                   "column": 2,
                   "header": "Generic_title",
                   "sheet": "Sheet_3",
                   "summary_column_list": column_dictionary3,
                   "offset": 0,
                   }},

    {"resp_dist": {"inicio_ensayo": 111, "fin_ensayo": 222, "respuesta": 333,
                   "bin_size": 1,
                   "bin_amount": 15,
                   }},
    ]
    """)


def create_sheets(workbook, sheets):
    """
    Se crea una lista con hojas de trabajo pertenecientes al directorio creado/abierto. Se pueden crear tantas listas
    como sea necesario. La lista tendrá el mismo orden que los argumentos de la función.\n
    :param sheets: Strings con los nombres que tendrá cada hoja de cálculo. La función admite una cantidad indefinida
de hojas.
    :param workbook: Archivo de tipo Workbook (creado mediante Openpyxl) en el que se generarán las hojas de trabajo.
    :return: Diccionario con listas asociadas al nombre dado como parámetro.
    """
    sheet_dict = {}
    for sheet in sheets:
        if sheet not in workbook.sheetnames:
            new_sheet = workbook.create_sheet(sheet)
        else:
            new_sheet = workbook[sheet]
        sheet_dict[sheet] = new_sheet
    return sheet_dict


class Analyzer:
    def __init__(self, fileName, temporaryDirectory, permanentDirectory, convertedDirectory, subjectList, suffix,
                 sheets, analysisList, timeColumn=None, markColumn=None, relocate=True):
        self.file_name = fileName
        self.temp_directory = temporaryDirectory
        self.perm_directory = permanentDirectory
        self.conv_directory = convertedDirectory
        self.subject_list = subjectList
        self.session_list = []
        self.suffix = suffix
        self.relocate = relocate
        self.sheets = sheets
        self.analysis_list = analysisList
        self.mark_column = markColumn
        self.time_column = timeColumn

    def get_sessions(self):
        """
        Se genera una lista temporal que contiene aquellos sujetos cuyos datos sí están en la carpeta temporal; después, se
        eliminan las columnas pertinentes de cada una de las listas de columnas. Los argumentos son: el directorio temporal
        en que están los datos brutos, una lista con los nombres de los sujetos, y las listas de columnas.\n
        """
        sujetosFaltantes = []
        dirTemp = sorted(listdir(self.temp_directory))
        listaTemp = []
        for sbj in self.subject_list:
            for datoTemp in dirTemp:
                if sbj == datoTemp.split('_')[0] and sbj not in listaTemp:
                    listaTemp.append(sbj)

        # Los sujetos que no forman parte de la lista temporal son agregados a la lista sujetosFaltantes para que sus
        # columnas sean eliminadas también del análisis.
        for sbj in self.subject_list:
            if sbj not in listaTemp:
                sujetosFaltantes.append(sbj)
        # Si faltan sujetos se imprime quiénes son. Si no, se indica con un mensaje.
        if len(sujetosFaltantes) == 0:
            print("All subjects have at least one session yet to analyze.")
        else:
            print(f"Missing subjects: {str(sujetosFaltantes)}")

        # Se hace una lista de listas con las sesiones presentes de cada sujeto.
        # El código compara el nombre de cada uno de los sujetos que sí tienen sesiones con el nombre de cada uno de los
        # archivos del directorio temporal. Si encuentra una coincidencia, toma el número de la sesión asociado con el
        # archivo encontrado y lo agrega a una sublista que contendrá todas las sesiones del sujeto en cuestión. El código
        # tolera sesiones salteadas
        indice = 0
        for sujetoPresente in listaTemp:
            self.session_list.append([])
            sublista = []
            for datoTemp in dirTemp:
                if sujetoPresente == datoTemp.split('_')[0]:
                    sublista.append(int(datoTemp.split('_')[-1]))
            self.session_list[indice] = sorted(sublista)
            indice += 1

        for i in range(len(listaTemp)):
            print(f"Sessions to analyze for subject {listaTemp[i]}: {self.session_list[i]}")

        # Se eliminan los elementos pertinentes de las listas de columnas si algún sujeto falta.
        for sujetoFaltante in sujetosFaltantes:
            #     for columnList in columnLists:
            #         if sujetoFaltante in subjectList:
            #             del columnList[subjectList.index(sujetoFaltante)]
            del self.subject_list[self.subject_list.index(sujetoFaltante)]
        print("\n")

    def converter(self):
        """
        Convierte archivos de texto plano de MedPC en hojas de cálculo en formato *.xlsx.
        Separa cada lista en dos columnas con base en el punto decimal.\n
        """
        for ind, sjt in enumerate(self.subject_list):
            for ssn in self.session_list[ind]:
                print(f"Converting session {ssn} of subject {sjt}.")
                # Pandas lee los datos y los escribe en el archivo convertido en 6 columnas separando por los espacios.
                # El argumento names indica cuántas columnas se crearán. Evita errores cuando se edita el archivo de Med.
                datos = pandas.read_csv(f"{self.temp_directory}{sjt}{self.suffix}{ssn}", header=None, names=range(6),
                                        sep=r'\s+')
                datos.to_excel(f"{self.conv_directory}{sjt}{self.suffix}{ssn}.xlsx", index=False, header=None)

                # Openpyxl abre el archivo creado por pandas, lee la hoja y la almacena en la variable hojaCompleta.
                archivoCompleto = load_workbook(f"{self.conv_directory}{sjt}{self.suffix}{ssn}.xlsx")
                hojaCompleta = archivoCompleto.active

                # Se genera una lista que contenga sub-listas con todos los valores de las listas dadas por Med.
                # Funciona para cualquier cantidad de listas.
                # Los datos se convierten en flotantes para que todos tengan punto decimal, y luego en string para que mas
                # adelante el método split los pueda separar por el punto.
                metalista = [[]]
                contadormetalista = 0

                columna1 = hojaCompleta['B']
                for fila in range(12, len(columna1) + 1):
                    for columna in range(2, 6 + 1):
                        if hojaCompleta[f"{get_column_letter(columna)}{fila}"].value is not None:
                            metalista[contadormetalista].append(str(float(hojaCompleta[get_column_letter(columna) +
                                                                                       str(fila)].value)))
                        elif hojaCompleta[get_column_letter(columna) + str(fila)].value is None and columna == 2:
                            metalista.append([])
                            contadormetalista += 1

                # Escribir cada sub-lista en una columna de excel separando por punto decimal.
                # Se utilizan expresiones regulares (regex) para indicar al programa que debe añadir ceros cuando pandas
                # los ha eliminado (cuando están al final de una cifra después de un punto decimal).
                regex1 = re.compile(r'^\d+\.\d{2}$')
                regex2 = re.compile(r'^\d+\.\d$')
                for ii in range(len(metalista)):
                    for j in range(len(metalista[ii])):
                        if regex1.search(metalista[ii][j]):
                            metalista[ii][j] += '0'
                        elif regex2.search(metalista[ii][j]):
                            metalista[ii][j] += '00'
                        hojaCompleta[get_column_letter((ii * 2) + 9) + str(j + 1)] = int(
                            metalista[ii][j].split('.')[0])
                        hojaCompleta[get_column_letter((ii * 2) + 10) + str(j + 1)] = int(
                            metalista[ii][j].split('.')[1])
                archivoCompleto.save(f"{self.conv_directory}{sjt}{self.suffix}{ssn}.xlsx")
                if self.relocate:
                    move(f"{self.temp_directory}{sjt}{self.suffix}{ssn}", f"{self.perm_directory}{sjt}{self.suffix}{ssn}")
            print('\n')

    def create_document(self):
        """
        Se inspecciona el directorio objetivo en busca de un archivo de resumen. Si existe, el archivo es abierto. Si no,
        es creado. Esta función debe asignarse a una variable.\n
        :return: Objeto de clase Workbook (Openpyxl).
        """
        if self.file_name in listdir(self.conv_directory):
            print('Summary file found. Opening...\n')
            wb = load_workbook(self.conv_directory + self.file_name)
        else:
            print('Summary file not found. Creating...\n')
            wb = Workbook()
        return wb

    def analyze(self, workbook, sheetDict):
        """
        Función principal de análisis. Toma los conteos realizados por las otras funciones y pega medidas de tendencia
        central en un archivo general de resumen. Además pega las listas completas en los archivos convertidos
        individuales.\n
        :param workbook: Hoja de cálculo generada por la función createDocument.
        :param sheetDict: Diccionario creado por la función create_sheets.
        """
        for index, subject in enumerate(self.subject_list):
            for session in self.session_list[index]:
                print(f"Trying session {session} of subject {subject}.")
                sujetoWb = load_workbook(f"{self.conv_directory}{subject}{self.suffix}{str(session)}.xlsx")
                sujetoWs = sujetoWb.worksheets[0]
                hojaind = sujetoWb.create_sheet('FullLists')
                tiempo = sujetoWs[self.time_column]
                marcadores = sujetoWs[self.mark_column]

                for analysis in self.analysis_list:
                    if "resp_dist" in analysis:
                        resp_dist_sheets = create_sheets(workbook, self.subject_list)
                        dist_indiv_sheet = sujetoWb.create_sheet('RespDistrib')

                    key, value = list(analysis.items())[0]

                    if key == "conteoresp":
                        respuestas_totales = []
                        for mark_index in range(1, value.get("measures", 1) + 1):
                            if mark_index == 1:
                                respuesta_parcial = conteoresp(marcadores, value["inicio_ensayo"],
                                                               value["fin_ensayo"],
                                                               value["respuesta"])
                            else:
                                respuesta_parcial = conteoresp(marcadores, value[f"inicio_ensyo{mark_index}"],
                                                               value[f"fin_ensayo{mark_index}"],
                                                               value[f"respuesta{mark_index}"])
                            respuestas_totales.extend(respuesta_parcial)
                        if value.get("substract", False):
                            respuestas_restadas = [resp - 1 if resp > 0 else resp for resp in respuestas_totales]
                            sheetDict[value["sheet"]][
                                get_column_letter(value["summary_column_list"][subject] + value.get("offset", 0)) + str(
                                    session + 3)] = mean(respuestas_restadas)
                            esccolumnas(hojaind, value["header"], value["column"], respuestas_restadas)
                        else:
                            sheetDict[value["sheet"]][
                                get_column_letter(value["summary_column_list"][subject] + value.get("offset", 0)) + str(
                                    session + 3)] = mean(respuestas_totales)
                            esccolumnas(hojaind, value["header"], value["column"], respuestas_totales)

                    elif key == "conteolat":
                        latencias_totales = []
                        for mark_index in range(1, value.get("measures", 1) + 1):
                            if mark_index == 1:
                                latencia_parcial = conteolat(marcadores, tiempo, value["inicio_ensayo"],
                                                             value["respuesta"])
                            else:
                                latencia_parcial = conteolat(marcadores, tiempo, value[f"inicio_ensayo{mark_index}"],
                                                             value[f"respuesta{mark_index}"])
                            latencias_totales.extend(latencia_parcial)
                        sheetDict[value["sheet"]][
                            get_column_letter(value["summary_column_list"][subject] + value.get("offset", 0)) + str(
                                session + 3)] = median(latencias_totales)
                        esccolumnas(hojaind, value["header"], value["column"], latencias_totales)

                    elif key == "conteototal":
                        respuestas_totales = 0
                        for mark_index in range(1, value.get("measures", 1) + 1):
                            if mark_index == 1:
                                respuesta_parcial = conteototal(marcadores, value["respuesta"])
                            else:
                                respuesta_parcial = conteototal(marcadores, value[f"respuesta{mark_index}"])
                            respuestas_totales += respuesta_parcial
                        sheetDict[value["sheet"]][
                            get_column_letter(value["summary_column_list"][subject] + value.get("offset", 0)) + str(
                                session + 3)] = respuestas_totales
                        esccolumnas(hojaind, value["header"], value["column"], [respuestas_totales])

                    elif key == "fetch":
                        cell_value = fetch(sujetoWs, value["cell_row"], value["cell_column"])
                        sheetDict[value["sheet"]][
                            get_column_letter(value["summary_column_list"][subject] + value.get("offset", 0)) + str(
                                session + 3)] = cell_value

                    elif key == "resp_dist":
                        superlist = resp_dist(marcadores, tiempo, trialStart=value["inicio_ensayo"],
                                              trialEnd=value["fin_ensayo"],
                                              response=value["respuesta"], bin_size=value["bin_size"],
                                              bin_amount=value["bin_amount"])
                        aggregated = []
                        means = []
                        for i in range(len(superlist[0])):
                            for sublist in superlist:
                                aggregated.append(sublist[i])
                            means.append(mean(aggregated))
                            aggregated = []
                        # Escribir en archivo de resumen
                        esccolumnas(resp_dist_sheets[subject], f"Session {session}", session + 1, means)
                        # Escribir en archivo individual
                        for ix, sublist in enumerate(superlist):
                            esccolumnas(dist_indiv_sheet, f"Trial {ix + 1}", ix + 1, sublist)

                sujetoWb.save(self.conv_directory + subject + self.suffix + str(session) + '.xlsx')
            print("\n")
        workbook.save(self.conv_directory + self.file_name)

    def convert(self):
        self.get_sessions()
        self.converter()

    def complete_analysis(self):
        self.get_sessions()
        self.converter()
        workbook = self.create_document()
        sheets = create_sheets(workbook, self.sheets)
        self.analyze(workbook, sheets)