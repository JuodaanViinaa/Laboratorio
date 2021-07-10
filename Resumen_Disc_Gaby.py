from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from os import listdir
from statistics import mean, median
import pandas
import re

archivo = 'Resumen_Disc_Py.xlsx'
# directorioBrutos = 'C:/Users/Admin/Desktop/Escape/Datos/Brutos/'
# directorioConvertidos = 'C:/Users/Admin/Desktop/Escape/Datos/ConvertidosPython/Escape/'
directorioBrutos = 'C:/Users/Admin/Documents/DISCRIMINABILITY2.0/SUBCHO/Brutos/'
directorioConvertidos = 'C:/Users/Admin/Documents/DISCRIMINABILITY2.0/SUBCHO/ConvertidosPy/'
sesionInicial = 16
sesionFinal = 16

sujetos = ['D1', 'D2', 'D3', 'D4', 'D5', 'D6', 'D7', 'D8']
columnasProp = [2, 3, 4, 5, 6, 7, 8, 9]
columnasResp = [2, 12, 22, 32, 42, 52, 62, 72] #pongo 10 columnas para agregar no. ensayos, elecDisc, elecNodidc y DisInd
columnasLatPal = [2, 7, 12, 17, 22, 27, 32, 37]
columnasLatCom = [2, 7, 12, 17, 22, 27, 32, 37]


# Convertidor
def convertir(columnas=6, subfijo=''):
    for ssn in range(sesionInicial, sesionFinal + 1):
        for sjt in range(len(sujetos)):
            print('Convirtiendo sesión ' + str(ssn) + ' de sujeto ' + sujetos[sjt] + '.')
            # Pandas lee los datos y los escribe en el archivo convertido en 6 columnas separando por los espacios.
            # El argumento names indica cuántas columnas se crearán. Evita errores cuando se edita el archivo de Med.
            datos = pandas.read_csv(directorioBrutos + sujetos[sjt] + subfijo + str(ssn), header=None,
                                    names=range(columnas), sep=r'\s+')
            datos.to_excel(directorioConvertidos + sujetos[sjt] + subfijo + str(ssn) + '.xlsx', index=False,
                           header=None)

            # Openpyxl abre el archivo creado por pandas, lee la hoja y la almacena en la variable hojaCompleta.
            archivoCompleto = load_workbook(directorioConvertidos + sujetos[sjt] + subfijo + str(ssn) + '.xlsx')
            hojaCompleta = archivoCompleto.active

            # Se genera una lista que contenga sub-listas con todos los valores de las listas dadas por Med.
            # Funciona para cualquier cantidad de listas.
            # Los datos se convierten en flotantes para que todos tengan punto decimal, y luego en string para que mas
            # adelante el método split los pueda separar por el punto.
            metalista = [[]]
            contadormetalista = 0

            columna1 = hojaCompleta['B']
            for fila in range(12, len(columna1)):
                for columna in range(2, columnas + 1):
                    if hojaCompleta[get_column_letter(columna) + str(fila)].value is not None:
                        metalista[contadormetalista].append(str(float(hojaCompleta[get_column_letter(columna) +
                                                                                   str(fila)].value)))
                    elif hojaCompleta[get_column_letter(columna) + str(fila)].value is None and columna == 2:
                        metalista.append([])
                        contadormetalista += 1

            # Escribir cada sub-lista en una columna de excel separando por punto decimal.
            # Se utilizan expresiones regulares (regex) para indicar al programa que debe añadir ceros cuando pandas
            # los ha eliminado (cuando están al final de una cifra después de un punto decimal).
            regex1 = re.compile(r'^\d+\.\d{2}$')
            for i in range(len(metalista)):
                for j in range(len(metalista[i])):
                    if regex1.search(metalista[i][j]):
                        metalista[i][j] += '0'
                    hojaCompleta[get_column_letter((i * 2) + 9) + str(j + 1)] = int(metalista[i][j].split('.')[0])
                    hojaCompleta[get_column_letter((i * 2) + 10) + str(j + 1)] = int(metalista[i][j].split('.')[1])
            archivoCompleto.save(directorioConvertidos + sujetos[sjt] + subfijo + str(ssn) + '.xlsx')
        print('\n')


# Función para crear hojas.
def hoja(nombre):
    if nombre not in wb.sheetnames:
        return wb.create_sheet(nombre)
    else:
        return wb[nombre]

# Función para contar respuestas por tipo de ensayo. Los argumentos son marcadores de Med.
def conteoresp(inicioEnsayo, finEnsayo, respuesta):
    contadorTemp = 0
    inicio = 0
    resp = []
    for n in range(1, len(marcadores)):
        if marcadores[n].value == inicioEnsayo:
            inicio = 1
        elif marcadores[n].value == respuesta and inicio == 1:
            contadorTemp += 1
        elif marcadores[n].value == finEnsayo and inicio == 1:
            inicio = 0
            resp.append(contadorTemp)
            contadorTemp = 0
        if len(resp) == 0:
             resp = [0]
    return resp

# Función para contar respuestas totales. El argumento es el marcador de Med.
def conteototal(respuesta):
    contador = 0
    for n in range(len(marcadores)):
        if marcadores[n].value == respuesta:
            contador += 1
    return contador

# Función para contar latencias. Si en un ensayo no hay respuestas que contar, la función resulta en una lista con un
# cero. Los argumentos son marcadores de Med.
def conteolat(inicioensayo, respuesta):
    inicio = 0
    lat = []
    tiempoini = 0
    for n in range(1, len(marcadores)):
        if marcadores[n].value == inicioensayo:
            inicio = 1
            tiempoini = tiempo[n].value
        elif marcadores[n].value == respuesta and inicio == 1:
            lat.append((tiempo[n].value - tiempoini) / 20)
            inicio = 0
    if len(lat) == 0:
        lat = [0]
    return lat

# Función para escribir listas en columnas. El argumento "restar" indica si se debe restar uno a las respuestas dadas
# en cada ensayo. Esto solo sucede en las palancas dado que se registra también la respuesta que le da inicio al ensayo.
def esccolumnas(titulo, columna, lista, restar):
    hojaind[get_column_letter(columna) + str(1)] = titulo
    if restar:
        for pos in range(len(lista)):
            if lista[pos] > 0:
                hojaind[get_column_letter(columna) + str(pos + 2)] = lista[pos] - 1
            else:
                hojaind[get_column_letter(columna) + str(pos + 2)] = lista[pos]
    else:
        for pos in range(len(lista)):
            hojaind[get_column_letter(columna) + str(pos + 2)] = lista[pos]

convertir(6,'_DISC_')

# Resumen
# Revisar si el archivo de resumen ya existe. De lo contrario, crearlo.
if archivo in listdir(directorioConvertidos):
    print('Archivo encontrado. Abriendo...')
    wb = load_workbook(directorioConvertidos + archivo)
else:
    print('Archivo no encontrado. Creando...')
    wb = Workbook()

# Crear todas las hojas.
proporciones = hoja('Proporciones')
respuestas = hoja('Respuestas')
latencias = hoja('Latencias')
comedero = hoja('Comedero')
latcomedero = hoja('LatComedero')


# Loop principal.
for sesion in range(sesionInicial, sesionFinal + 1):
    print('\nIntentando sesión ' + str(sesion) + '...')
    for sujeto in range(len(sujetos)):
        print('Intentando sujeto ' + sujetos[sujeto] + '...')
        sujetoWb = load_workbook(directorioConvertidos + sujetos[sujeto] + '_DISC_' + str(sesion) + '.xlsx')
        sujetoWs = sujetoWb.worksheets[0]
        tiempo = sujetoWs['M']
        marcadores = sujetoWs['N']

        # Abrir o crear la hoja para pegar respuestas individuales.
        if 'Respuestas por ensayo' not in sujetoWb.sheetnames:
            hojaind = sujetoWb.create_sheet('Respuestas por ensayo')
        else:
            hojaind = sujetoWb['Respuestas por ensayo']

        # Proporciones
        proporciones[get_column_letter(columnasProp[sujeto]) + str(sesion + 3)] = sujetoWs.cell(14, 6).value
        # Se toma el valor de la celda F14, que corresponde a los valores de fila 14 y columna 6 debido a que
        # el conteo comienza desde 0 y no desde 1.
        respuestas[get_column_letter(columnasResp[sujeto] + 4) + str(sesion + 3)] = sujetoWs.cell(15, 2).value #ensayos
        respuestas[get_column_letter(columnasResp[sujeto] + 5) + str(sesion + 3)] = sujetoWs.cell(12, 6).value #elecdisc
        respuestas[get_column_letter(columnasResp[sujeto] + 6) + str(sesion + 3)] = sujetoWs.cell(13, 6).value #elecNoD
        # Conteo de respuestas en palancas en ensayos forzados.
        # Se escriben las respuestas por ensayo en el archivo individual.
        # Se resta una unidad a las medias debido a que conteoresp() cuenta también la respuesta que inicia el ensayo.
        resPalForzDiscBaj = conteoresp(114, 116, 902)
        mediaResPalForzDiscBaj = mean(resPalForzDiscBaj) + 1
        esccolumnas('PalForzDiscBaj', 1, resPalForzDiscBaj, False)
        # 114: Inicio TL Forz Disc Baj  //  180: Fin ensayo (general)  //  201: Resp Pal Disc
        resPalForzDiscAlt = conteoresp(115, 117, 904)
        mediaResPalForzDiscAlt = mean(resPalForzDiscAlt) + 1
        esccolumnas('PalForzDiscAlt', 3, resPalForzDiscAlt, False)
        # 115: Inicio TL Forz Disc NoRef  //  180: Fin ensayo (general)  //  201: Res Pal Disc
        resPalForzNoDiscBaj = conteoresp(134, 140, 906)
        mediaResPalForzNoDiscBaj = mean(resPalForzNoDiscBaj)+ 1
        esccolumnas('PalForzNoDiscBaj', 5, resPalForzNoDiscBaj, False)
        # 134: Inicio TL Forz NoDisc 1  //  180: Fin ensayo  //  202: Resp Pal NoDisc
        resPalForzNoDiscAlt = conteoresp(137, 143, 908)
        mediaResPalForzNoDiscAlt = mean(resPalForzNoDiscAlt) + 1
        esccolumnas('PalForzNoDiscAlt', 7, resPalForzNoDiscAlt, False)
        # 137: Inicio TL Forz NoDisc 2  //  180: Fin ensayo  //  202: Resp Pal NoDisc
        respuestas[get_column_letter(columnasResp[sujeto]) + str(sesion + 3)] = mediaResPalForzDiscBaj
        respuestas[get_column_letter(columnasResp[sujeto] + 1) + str(sesion + 3)] = mediaResPalForzDiscAlt
        respuestas[get_column_letter(columnasResp[sujeto] + 2) + str(sesion + 3)] = mediaResPalForzNoDiscBaj
        respuestas[get_column_letter(columnasResp[sujeto] + 3) + str(sesion + 3)] = mediaResPalForzNoDiscAlt

        # Conteo de latencias a palancas en eslabones iniciales.
        # Las latencias por ensayo se pegan en el archivo individual.
        latPalDisc = conteolat(112, 113)
        medianaLatPalDisc = median(latPalDisc)
        esccolumnas('LatPalDisc', 9, latPalDisc, False)
        # 112: IL Forz Disc  //  113: Res Pal
        latPalNoDisc = conteolat(132, 133)
        medianaLatPalNoDisc = median(latPalNoDisc)
        esccolumnas('LatPalNoDisc', 11, latPalNoDisc, False)
        # 132: IL Forzado NoDisc  //  133: Res Pal
        latencias[get_column_letter(columnasLatPal[sujeto]) + str(sesion + 3)] = medianaLatPalDisc
        latencias[get_column_letter(columnasLatPal[sujeto] + 1) + str(sesion + 3)] = medianaLatPalNoDisc

        # Conteo de respuestas en comederos en ensayos forzados
        # Las respuestas por ensayo se pegan en el archivo individual.
        resComForzDiscBaj = conteoresp(114, 116, 203)
        mediaResComForzDiscBaj = mean(resComForzDiscBaj)
        esccolumnas('ComForzDiscBaj', 13, resComForzDiscBaj, False)
        # 114: Inicio TL Forz Disc Ref  //  16: Fin ensayo  //  203: Res Com
        resComForzDiscAlt = conteoresp(115, 117, 203)
        mediaResComForzDiscAlt = mean(resComForzDiscAlt)
        esccolumnas('ComForzDiscAlt', 15, resComForzDiscAlt, False)
        # 115: Inicio TL Forz Disc NoRef  //  117: Fin ensayo (por TF)  //  203: Res Com
        resComForzNoDiscBaj = conteoresp(134, 183, 203)
        mediaResComForzNoDiscBaj = mean(resComForzNoDiscBaj)
        esccolumnas('ComForzNoDiscBaj', 17, resComForzNoDiscBaj, False)
        # 134: Inicio TL Forz NoDisc 1  //  40: Fin ensayo  //  203: Res Com
        resComForzNoDiscAlt = conteoresp(137, 184, 203)
        mediaResComForzNoDiscAlt = mean(resComForzNoDiscAlt)
        esccolumnas('ComForzNoDiscAlt', 19, resComForzNoDiscAlt, False)
        # 137: Inicio TL Forz NoDisc 2  //  43: Fin ensayo  //  203: Res Com
        comedero[get_column_letter(columnasResp[sujeto]) + str(sesion + 3)] = mediaResComForzDiscBaj
        comedero[get_column_letter(columnasResp[sujeto] + 1) + str(sesion + 3)] = mediaResComForzDiscAlt
        comedero[get_column_letter(columnasResp[sujeto] + 2) + str(sesion + 3)] = mediaResComForzNoDiscBaj
        comedero[get_column_letter(columnasResp[sujeto] + 3) + str(sesion + 3)] = mediaResComForzNoDiscAlt

        # Conteo de latencias a palancas en eslabones iniciales.
        # Las latencias por ensayo se pegan en el archivo individual.
        latComDisc = conteolat(112, 203)
        medianaLatComDisc = median(latPalDisc)
        esccolumnas('LatPalDisc', 9, latComDisc, False)
        # 112: IL Forz Disc  //  113: Res Pal
        latComNoDisc = conteolat(132, 203)
        medianaLatComNoDisc = median(latComNoDisc)
        esccolumnas('LatPalNoDisc', 11, latComNoDisc, False)
        # 132: IL Forzado NoDisc  //  133: Res Pal
        latencias[get_column_letter(columnasLatCom[sujeto]) + str(sesion + 3)] = medianaLatComDisc
        latencias[get_column_letter(columnasLatCom[sujeto] + 1) + str(sesion + 3)] = medianaLatComNoDisc

        sujetoWb.save(directorioConvertidos + sujetos[sujeto] + '_DISC_' + str(sesion) + '.xlsx')

wb.save(directorioConvertidos + archivo)