from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from os import listdir
from statistics import mean, median
import pandas
import re
from shutil import move

archivo = 'ResumenAlter.xlsx'
# directorioBrutos = 'C:/Users/Admin/Desktop/Escape/Datos/Brutos/'
# directorioConvertidos = 'C:/Users/Admin/Desktop/Escape/Datos/ConvertidosPython/Escape/'
directorioTemporal = '/home/daniel/Documents/Doctorado/Proyecto de Doctorado/ExperimentoEscape/Temporal/'
directorioBrutos = '/home/daniel/Documents/Doctorado/Proyecto de Doctorado/ExperimentoEscape/AlterBrutos/'
directorioConvertidos = '/home/daniel/Documents/Doctorado/Proyecto de Doctorado/ExperimentoEscape/AlterConvertidos/'
sesionesPresentes = []

sujetos = ['E3', 'E4', 'E5', 'E7', 'E8', 'E9', 'E1']
sujetosFaltantes = []
cols = [2, 7, 12, 17, 22, 27, 32]

# Se genera una lista temporal que contiene aquellos sujetos cuyos datos sí están en la carpeta temporal.
dirTemp = sorted(listdir(directorioTemporal))
listaTemp = []
for sujeto in sujetos:
    for datoTemp in dirTemp:
        if sujeto == datoTemp.split('_')[0] and sujeto not in listaTemp:
            listaTemp.append(sujeto)

# Los sujetos que no forman parte de la lista temporal son agregados a la lista sujetosFaltantes para que sus
# columnas sean eliminadas también del análisis.
for sujeto in sujetos:
    if sujeto not in listaTemp:
        sujetosFaltantes.append(sujeto)
# Si faltan sujetos se imprime quiénes son. Si no, se indica con un mensaje.
if len(sujetosFaltantes) == 0:
    print('Todos los sujetos tienen al menos una sesión por analizar.')
else:
    print('Sujetos faltantes: ' + str(sujetosFaltantes))

# Se hace una lista de listas con las sesiones presentes de cada sujeto.
indice = 0
for sujetoPresente in listaTemp:
    sesionesPresentes.append([])
    sublista = []
    for datoTemp in dirTemp:
        if sujetoPresente == datoTemp.split('_')[0]:
            sublista.append(int(datoTemp.split('_')[-1]))
    sesionesPresentes[indice] = sorted(sublista)
    indice += 1

for i in range(len(listaTemp)):
    print('Sesiones presentes del sujeto ' + listaTemp[i] + ': ' + str(sesionesPresentes[i]))
print('\n')

# Se eliminan los elementos pertinentes de las listas de columnas si algún sujeto falta.
for sujetoFaltante in sujetosFaltantes:
    if sujetoFaltante in sujetos:
        del cols[sujetos.index(sujetoFaltante)]
        del sujetos[sujetos.index(sujetoFaltante)]


# Convertidor
def convertir(columnas=6, subfijo=''):
    for sjt in range(len(sujetos)):
        for ssn in sesionesPresentes[sjt]:
            print('Convirtiendo sesión ' + str(ssn) + ' de sujeto ' + sujetos[sjt] + '.')
            # Pandas lee los datos y los escribe en el archivo convertido en 6 columnas separando por los espacios.
            # El argumento names indica cuántas columnas se crearán. Evita errores cuando se edita el archivo de Med.
            datos = pandas.read_csv(directorioTemporal + sujetos[sjt] + subfijo + str(ssn), header=None,
                                    names=range(columnas), sep=r'\s+')
            datos.to_excel(directorioConvertidos + sujetos[sjt] + subfijo + str(ssn) + '.xlsx', index=False,
                           header=None)

            # Openpyxl abre el archivo creado por pandas, lee la hoja y la almacena en la variable hojaCompleta.
            archivoCompleto = load_workbook(directorioConvertidos + sujetos[sjt] + subfijo + str(ssn) + '.xlsx')
            hojaCompleta = archivoCompleto.active

            # Se genera una lista que contenga sub-listas con todos los valores de las listas dadas por Med.
            # Funciona para cualquier cantidad de listas.
            # Los datos se convierten en flotantes para que todos tengan punto decimal, y luego en string para que mas
            # adelante el método split los pueda separar por el punto.
            metalista = [[]]
            contadormetalista = 0

            columna1 = hojaCompleta['B']
            for fila in range(12, len(columna1) + 1):
                for columna in range(2, columnas + 1):
                    if hojaCompleta[get_column_letter(columna) + str(fila)].value is not None:
                        metalista[contadormetalista].append(str(float(hojaCompleta[get_column_letter(columna) +
                                                                                   str(fila)].value)))
                    elif hojaCompleta[get_column_letter(columna) + str(fila)].value is None and columna == 2:
                        metalista.append([])
                        contadormetalista += 1

            # Escribir cada sub-lista en una columna de excel separando por punto decimal.
            # Se utilizan expresiones regulares (regex) para indicar al programa que debe añadir ceros cuando pandas
            # los ha eliminado (cuando están al final de una cifra después de un punto decimal).
            regex1 = re.compile(r'^\d+\.\d{2}$')
            for ii in range(len(metalista)):
                for j in range(len(metalista[ii])):
                    if regex1.search(metalista[ii][j]):
                        metalista[ii][j] += '0'
                    hojaCompleta[get_column_letter((ii * 2) + 9) + str(j + 1)] = int(metalista[ii][j].split('.')[0])
                    hojaCompleta[get_column_letter((ii * 2) + 10) + str(j + 1)] = int(metalista[ii][j].split('.')[1])
            archivoCompleto.save(directorioConvertidos + sujetos[sjt] + subfijo + str(ssn) + '.xlsx')
            move(directorioTemporal + sujetos[sjt] + subfijo + str(ssn), directorioBrutos + sujetos[sjt] + subfijo + str(ssn))
        print('\n')


# Función para crear hojas.
def hoja(nombre):
    if nombre not in wb.sheetnames:
        return wb.create_sheet(nombre)
    else:
        return wb[nombre]


# Función para contar latencias. Si en un ensayo no hay respuestas que contar, la función resulta en una lista con un
# cero. Los argumentos son marcadores de Med.
def conteolat(inicioensayo, respuesta):
    inicio = 0
    lat = []
    tiempoini = 0
    for n in range(1, len(marcadores)):
        if marcadores[n].value == inicioensayo:
            inicio = 1
            tiempoini = tiempo[n].value
        elif marcadores[n].value == respuesta and inicio == 1:
            lat.append((tiempo[n].value - tiempoini) / 20)
            inicio = 0
    if len(lat) == 0:
        lat = [0]
    return lat


# Función para escribir listas en columnas. El argumento "restar" indica si se debe restar uno a las respuestas dadas
# en cada ensayo. Esto solo sucede en las palancas dado que se registra también la respuesta que le da inicio al ensayo.
def esccolumnas(titulo, columna, lista, restar):
    hojaind[get_column_letter(columna) + str(1)] = titulo
    if restar:
        for pos in range(len(lista)):
            if lista[pos] > 0:
                hojaind[get_column_letter(columna) + str(pos + 2)] = lista[pos] - 1
            else:
                hojaind[get_column_letter(columna) + str(pos + 2)] = lista[pos]
    else:
        for pos in range(len(lista)):
            hojaind[get_column_letter(columna) + str(pos + 2)] = lista[pos]


convertir(subfijo='_ALTER_')

# Resumen
# Revisar si el archivo de resumen ya existe. De lo contrario, crearlo.
if archivo in listdir(directorioConvertidos):
    print('Archivo encontrado. Abriendo...')
    wb = load_workbook(directorioConvertidos + archivo)
else:
    print('Archivo no encontrado. Creando...')
    wb = Workbook()

# Crear todas las hojas.
latencias = hoja('Latencias')

# Loop principal.
for sujeto in range(len(sujetos)):
    print('\nIntentando sujeto ' + sujetos[sujeto] + '...')
    for sesion in sesionesPresentes[sujeto]:
        print('Intentando sesión ' + str(sesion) + '...')
        sujetoWb = load_workbook(directorioConvertidos + sujetos[sujeto] + '_ALTER_' + str(sesion) + '.xlsx')
        sujetoWs = sujetoWb.worksheets[0]
        tiempo = sujetoWs['K']
        marcadores = sujetoWs['L']

        # Abrir o crear la hoja para pegar respuestas individuales.
        if 'Latencias por ensayo' not in sujetoWb.sheetnames:
            hojaind = sujetoWb.create_sheet('Latencias por ensayo')
        else:
            hojaind = sujetoWb['Latencias por ensayo']

        # Conteo de latencias a todos los estímulos.
        # Las latencias por ensayo se pegan en el archivo individual.
        # Latencias a palanca izquierda
        latPalDiscBlanco = conteolat(202, 303)
        esccolumnas('LatPalDiscBlanco', 1, latPalDiscBlanco, False)

        latPalDiscRojo = conteolat(202, 306)
        esccolumnas('LatPalDiscRojo', 3, latPalDiscRojo, False)

        latPalDiscAzul = conteolat(202, 309)
        esccolumnas('LatPalDiscAzul', 5, latPalDiscAzul, False)

        latPalIzqTotal = latPalDiscBlanco + latPalDiscRojo + latPalDiscAzul
        medianaLatPalIzq = median(latPalIzqTotal)
        # 202: Inicio ensayo izquierdo; 303, 306, 309: Respuestas en pal izq ante cada estímulo.

        # Latencias a palanca derecha
        latPalNoDiscBlanco = conteolat(206, 312)
        esccolumnas('LatPalNoDiscBlanco', 7, latPalNoDiscBlanco, False)

        latPalNoDiscRojo = conteolat(206, 315)
        esccolumnas('LatPalNoDiscRojo', 9, latPalNoDiscRojo, False)

        latPalNoDiscAzul = conteolat(206, 318)
        esccolumnas('LatPalNoDiscAzul', 11, latPalNoDiscAzul, False)

        latPalDerTotal = latPalNoDiscBlanco + latPalNoDiscRojo + latPalNoDiscAzul
        medianaLatPalDer = median(latPalDerTotal)
        # 206: Inicio ensayo derecho; 312, 315, 318: Respuestas en pal der ante cada estímulo.

        # Latencias a nosepoke
        latNosepoke = conteolat(210, 321)
        esccolumnas('LatNosepoke', 13, latNosepoke, False)
        medianaLatNosepoke = median(latNosepoke)

        latencias[get_column_letter(cols[sujeto]) + str(sesion + 3)] = medianaLatPalIzq
        latencias[get_column_letter(cols[sujeto] + 1) + str(sesion + 3)] = medianaLatPalDer
        latencias[get_column_letter(cols[sujeto] + 2) + str(sesion + 3)] = medianaLatNosepoke

        sujetoWb.save(directorioConvertidos + sujetos[sujeto] + '_ALTER_' + str(sesion) + '.xlsx')

wb.save(directorioConvertidos + archivo)
