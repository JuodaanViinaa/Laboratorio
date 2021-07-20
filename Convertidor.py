from os import listdir
import pandas
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re
from shutil import move

directorioTemporal = '/home/daniel/Documents/Doctorado/Proyecto de Doctorado/ExperimentoEscape/Temporal/'
directorioBrutos = '/home/daniel/Documents/Doctorado/Proyecto de Doctorado/ExperimentoEscape/PruebasBrutos/'
directorioConvertidos = '/home/daniel/Documents/Doctorado/Proyecto de Doctorado/ExperimentoEscape/Flex/'
sujetos = ['E3', 'E4', 'E5', 'E7', 'E8', 'E9']
sujetosFaltantes = []
sesionesIniciales = []
sesionesFinales = []

# Se genera una lista temporal que contiene aquellos sujetos cuyos datos sí están en la carpeta temporal.
dirTemp = sorted(listdir(directorioTemporal))
listaTemp = []
for sujeto in sujetos:
    for datoTemp in dirTemp:
        if sujeto == datoTemp.split('_')[0] and sujeto not in listaTemp:
            listaTemp.append(sujeto)

# Los sujetos que no forman parte de la lista temporal son agregados a la lista sujetosFaltantes para que sus
# columnas sean eliminadas también del análisis.
for sujeto in sujetos:
    if sujeto not in listaTemp:
        sujetosFaltantes.append(sujeto)
# Si faltan sujetos se imprime quiénes son. Si no, se indica con un mensaje.
if len(sujetosFaltantes) == 0:
    print('Todos los sujetos tienen al menos una sesión por analizar.')
else:
    print('Sujetos faltantes: ' + str(sujetosFaltantes))

# Se hace una lista con las sesiones presentes para cada sujeto. Después se toma el valor más bajo y el más alto
# y se escriben en las listas de sesionesIniciales y sesionesFinales en la posición perteneciente a cada sujeto.
for sujetoPresente in listaTemp:
    sesionesPresentes = []
    for datoTemp in dirTemp:
        if sujetoPresente == datoTemp.split('_')[0]:
            sesionesPresentes.append(datoTemp.split('_')[-1])
    sesionesIniciales.append(int(min(sesionesPresentes)))
    sesionesFinales.append(int(max(sesionesPresentes)))
for i in range(len(listaTemp)):
    print('Sesiones inicial y final del sujeto ' + listaTemp[i] + ': ' + str(sesionesIniciales[i]) + ', ' + str(sesionesFinales[i]))
print('\n')

# La lista de sujetos se iguala a listaTemp, que contiene solo los sujetos presentes
sujetos = listaTemp


def convertir(columnas=6, subfijo=''):
    for sjt in range(len(sujetos)):
        for ssn in range(sesionesIniciales[sjt], sesionesFinales[sjt] + 1):
            print('Convirtiendo sesión ' + str(ssn) + ' de sujeto ' + sujetos[sjt] + '.')
            # Pandas lee los datos y los escribe en el archivo convertido en 6 columnas separando por los espacios.
            # El argumento names indica cuántas columnas se crearán. Evita errores cuando se edita el archivo de Med.
            datos = pandas.read_csv(directorioTemporal + sujetos[sjt] + subfijo + str(ssn), header=None,
                                    names=range(columnas), sep=r'\s+')
            datos.to_excel(directorioConvertidos + sujetos[sjt] + subfijo + str(ssn) + '.xlsx', index=False,
                           header=None)

            # Openpyxl abre el archivo creado por pandas, lee la hoja y la almacena en la variable hojaCompleta.
            archivoCompleto = load_workbook(directorioConvertidos + sujetos[sjt] + subfijo + str(ssn) + '.xlsx')
            hojaCompleta = archivoCompleto.active

            # Se genera una lista que contenga sub-listas con todos los valores de las listas dadas por Med.
            # Funciona para cualquier cantidad de listas.
            # Los datos se convierten en flotantes para que todos tengan punto decimal, y luego en string para que mas
            # adelante el método split los pueda separar por el punto.
            metalista = [[]]
            contadormetalista = 0

            columna1 = hojaCompleta['B']
            for fila in range(12, len(columna1)):
                for columna in range(2, columnas + 1):
                    if hojaCompleta[get_column_letter(columna) + str(fila)].value is not None:
                        metalista[contadormetalista].append(str(float(hojaCompleta[get_column_letter(columna) +
                                                                                   str(fila)].value)))
                    elif hojaCompleta[get_column_letter(columna) + str(fila)].value is None and columna == 2:
                        metalista.append([])
                        contadormetalista += 1

            # Escribir cada sub-lista en una columna de excel separando por punto decimal.
            # Se utilizan expresiones regulares (regex) para indicar al programa que debe añadir ceros cuando pandas
            # los ha eliminado (cuando están al final de una cifra después de un punto decimal).
            regex1 = re.compile(r'^\d+\.\d{2}$')
            for i in range(len(metalista)):
                for j in range(len(metalista[i])):
                    if regex1.search(metalista[i][j]):
                        metalista[i][j] += '0'
                    hojaCompleta[get_column_letter((i * 2) + 9) + str(j + 1)] = int(metalista[i][j].split('.')[0])
                    hojaCompleta[get_column_letter((i * 2) + 10) + str(j + 1)] = int(metalista[i][j].split('.')[1])
            archivoCompleto.save(directorioConvertidos + sujetos[sjt] + subfijo + str(ssn) + '.xlsx')
            move(directorioTemporal + sujetos[sjt] + subfijo + str(ssn), directorioBrutos + sujetos[sjt] + subfijo + str(ssn))
        print('\n')


convertir(subfijo='_ESCAPE_')
